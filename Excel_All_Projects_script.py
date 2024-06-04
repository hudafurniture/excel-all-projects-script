#--------------------------------
#This script Adds all Ardis .R41 Projects in the folder where script file is locvated to an excel workbook named "Projects_Log",
#It adds Project Name and it's created and modified time each into new row
#If project already exist it will update modified time only of the project in that row
#The Workbook divided into sheets where every sheet is a diffirent year
#--------------------------------

import os
import openpyxl
from openpyxl import load_workbook
from datetime import datetime


def main():

    workbook_name = 'Projects_Log.xlsx'
    wb = load_workbook(workbook_name)
    page = wb["2021"]
    pageInt =  int("".join(filter(str.isdigit, str(page))))

    filePath = os.getcwd()
    toAddLst = []
    
    for root, dirs, files in os.walk(filePath):
        for file in files:
            if file.endswith(".R41"):
                file=os.path.join(root, file)
                projectName = os.path.splitext(os.path.basename(file))[0]
                createdTime = datetime.fromtimestamp(os.path.getctime(file))
                createdDate = createdTime.strftime("%d-%m-%Y %H:%M")
                modifiedTime = datetime.fromtimestamp(os.path.getmtime(file))
                modifiedDate = modifiedTime.strftime("%d-%m-%Y %H:%M")
                modifiedYear = modifiedTime.year
           
                # New data to write:
                if (modifiedYear == pageInt):
                    toAddLst.append([projectName,modifiedDate])

    for info in toAddLst:
        projectExist = findProject(info, page)
        if (projectExist[0] == 0):   #if project doesn't exist
            page.append(info)      
            
    wb.save(filename=workbook_name)
    
    
            
 
 
 


def findProject(projecData, page):
    for i in range(1,page.max_row):
        if page.cell(row=i, column=1).value == projecData[0]:
            #print("exist")
            return [i, projecData[0]]     
    return [0,""]  #project was not found


















if __name__ == "__main__":
    main()


